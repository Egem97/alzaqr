import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import io
import re
import json
from datetime import datetime
from styles import styles
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode ,JsCode
from io import StringIO, BytesIO
from utils.helpers import crear_pdf, generar_qr, crear_pdf_qr_tunel, crear_pdf_packing_list
from utils.components import aggrid_builder,aggrid_builder_prod,aggrid_editing_prod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_dni_datos(text):
    """
    Extrae DNI (números) y datos (texto) del texto
    """
    dni = ""
    datos = ""
    
    if pd.isna(text) or text == "":
        return dni, datos
    
    text = str(text)
    
    # Extraer todos los números de 7-8 dígitos (DNI)
    dni_pattern = r'\b(\d{7,8})\b'
    dni_matches = re.findall(dni_pattern, text)
    
    if dni_matches:
        dni = dni_matches[0]  # Tomar el primer DNI encontrado
    
    # Extraer todo el texto, excluyendo los números de DNI
    # Remover los DNIs encontrados del texto
    for dni_match in dni_matches:
        text = text.replace(dni_match, '')
    
    # Limpiar el texto restante
    # Remover caracteres de control y caracteres extraños
    datos = re.sub(r'[^\x20-\x7E]', '', text)
    # Remover comillas múltiples y caracteres JSON
    datos = re.sub(r'["\']+', '', datos)
    # Remover caracteres como :, {, }, [, ]
    datos = re.sub(r'[{}[\]:,]+', ' ', datos)
    # Remover palabras específicas: Dni, Nombres, Apellidos (insensible a mayúsculas/minúsculas)
    datos = re.sub(r'\b(Dni|Nombres|Apellidos)\b', '', datos, flags=re.IGNORECASE)
    # Remover espacios múltiples
    datos = re.sub(r'\s+', ' ', datos)
    # Limpiar espacios al inicio y final
    datos = datos.strip()
    
    return dni, datos

def procesar_columna_text(df):
    """
    Procesa la columna text del DataFrame y crea nuevas columnas para DNI y datos
    """
    # Aplicar la función de extracción a cada fila
    resultados = df['text'].apply(extract_dni_datos)
    
    # Crear las nuevas columnas
    df['DNI'] = [result[0] for result in resultados]
    df['DATOS'] = [result[1] for result in resultados]
    
    return df

def crear_excel_corporativo(df, nombre_archivo="datos_procesados.xlsx"):
    """
    Crea un archivo Excel con formato corporativo
    """
    # Crear un nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Procesados"
    
    # Definir estilos corporativos
    # Color de encabezado (azul corporativo)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Color de datos alternados
    row_fill_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Fuente para datos
    data_font = Font(size=11)
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alineación
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Agregar título
    ws['A1'] = "REGISTRO DE MARCACIÓN GARITA"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:F1')
    
    # Agregar fecha de generación
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = left_alignment
    ws.merge_cells('A2:F2')
    
    # Agregar datos del DataFrame
    # Escribir encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Escribir datos
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            
            # Aplicar color alternado a las filas
            if row_idx % 2 == 0:
                cell.fill = row_fill_1
            else:
                cell.fill = row_fill_2
            
            # Alineación específica por columna
            if col_idx in [1, 2]:  # Fecha y Hora
                cell.alignment = center_alignment
            elif col_idx == 3:  # Formato
                cell.alignment = center_alignment
            elif col_idx == 4:  # DNI
                cell.alignment = center_alignment
            else:  # DATOS
                cell.alignment = left_alignment
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15,  # Fecha
        'B': 10,  # Hora
        'C': 12,  # Formato
        'D': 12,  # DNI
        'E': 40,  # DATOS
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Agregar estadísticas al final
    stats_row = len(df) + 6
    ws.cell(row=stats_row, column=1, value="ESTADÍSTICAS")
    ws.cell(row=stats_row, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(f'A{stats_row}:F{stats_row}')
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value=f"Total de registros: {len(df)}")
    ws.cell(row=stats_row, column=3, value=f"DNIs extraídos: {df['DNI'].notna().sum()}")
    ws.cell(row=stats_row, column=5, value=f"Registros con datos: {df['DATOS'].notna().sum()}")
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def gestion_humana_packing():
    styles(2)
    
    
    col_head1,col_head2, = st.columns([6,6])
    with col_head1:
        st.title("📄GH Tool")
    with col_head2:
        uploaded_file = st.file_uploader("Escoja su archivo csv", accept_multiple_files=False,type=['csv'],key="uploaded_file")
    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file, sep=",", quotechar='"')
       
        # Si todavía sale como una sola columna, volver a separarla manualmente
        if df.shape[1] == 1:
            df = df.iloc[:, 0].str.split(",", expand=True)
            df = df.drop(columns=[7, 8,9, 10, 11])
            df[6] = df[6].str.replace('"0"','')
            df["text"] = df[4] + "" + df[5] + "" + df[6]
            df = df.drop(columns=[4, 5,6])
            df.columns = ['date', 'time', 'time_zone', 'format', 'text']
            df["time"] = df["time"].str.replace('"','')
            df["time_zone"] = df["time_zone"].str.replace('"','')
            df["format"] = df["format"].str.replace('"','')
            #df["text"] = df["text"].str.replace('"','')
            #df['text'] = df['text'].apply(extract_dni_datos)
        else:
            df = df.drop(columns=['notes', 'favorite','date_utc', 'time_utc', 'metadata'])
        
        
        #['date', 'time', 'time_zone', 'format', 'text']
        # Procesar la columna text para extraer DNI, nombres y apellidos
        df_procesado = procesar_columna_text(df)
        df_procesado = df_procesado.drop(columns=['text','time_zone'])
        df_procesado = df_procesado.rename(columns={"date":"Fecha","time":"Hora","format":"Formato"})
        df_procesado["DATOS"] = df_procesado["DATOS"].str.replace("http //", "--")
        df_procesado["DATOS"] = df_procesado["DATOS"].str.replace("Apellidos", "")
        st.dataframe(df_procesado,hide_index=True)
        
        # Crear Excel con formato corporativo
        excel_file = crear_excel_corporativo(df_procesado)
        st.download_button(
                label="📥 Descargar Excel",
                data=excel_file.getvalue(),
                file_name="reporte_marcaciones_gh.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        